from openpyxl import workbook

wb = workbook.Workbook()
sheet = wb.active
sheet.title = "工作表1"  ##更改工作表名稱
sheet.append([1, 2, 3])  ##寫入資料
sheet.cell(1, 1).fill = "FF0000"  ##更改儲存格顏色
# sheet.cell(row=1, column=1).value = "Hello World"  ##寫入資料

# cell = sheet.cell(row=1, column=1)  ##讀取資料


wb.save("xxx.xlsx")  ##儲存

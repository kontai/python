from openpyxl import load_workbook

#read excel file
wb=load_workbook("../files/p1.xlsx",read_only=True)


#sheet的相闗操作

#獲取所有sheet
print(wb.sheetnames)

#選擇sheet
sheet=wb[wb.sheetnames[0]]

#選取單元格
cell=sheet.cell(row=1,column=1)
print(cell.value)

#讀取所有sheet第一行第一列內容
#1
for sheet in wb:
    cell=sheet.cell(1,1)
    print(cell.value)
#2
for sheet in wb.worksheets:
    cell=sheet.cell(1,1)
    print(cell.value)
#3
for name in wb.sheetnames:
    sheet=wb[name]
    cell=sheet.cell(1,1)
    print(cell.value)

e=enumerate(wb.sheetnames)
for i,name in e:
    sheet=wb[name]
    cell=sheet.cell(1,1)
    print(f'sheet{i}:{cell.value}')
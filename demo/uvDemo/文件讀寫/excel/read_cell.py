import io
import sys

from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="UTF-8")  # 解決中文亂碼問題

wb = load_workbook("../files/p1.xlsx", data_only=True)
# sheet = wb.active  # 獲取目前開啟的工作表(較安全)
# sheet = wb[wb.sheetnames[0]]
sheet = wb[wb.sheetnames[1]]  ##獲取第一個工作表的名稱
cell = sheet.cell(1, 1)

"""
print(cell.value)  ##獲取單元格內容
print(cell.style)  ##獲取單元格樣式
print(cell.alignment)  ##對齊方式
print(cell.font)  ##字體
"""

# 讀取單元格的值
# print(sheet["A1"].value)

##獲取第N行的所有單元格
print(sheet[1])

##獲取所有行
# print([r[2] for r in sheet.rows])

##獲取所有列
print([c[2] for c in sheet.columns])
